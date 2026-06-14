import os
import time
import shutil
import hashlib
import openpyxl
from server import PromptServer
from aiohttp import web

# ── Paths ────────────────────────────────────────────────────────────────────

def get_source_path():
    """Returns the absolute path to Styles.xlsx inside this node's folder."""
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), "Styles.xlsx")

def get_shadow_path():
    """Returns the absolute path to the shadow copy."""
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), "._styles_cache.xlsx")


# ── Cache State ──────────────────────────────────────────────────────────────

_EXCEL_CACHE = None          # {tab_name: {style_name: [pos, neg], ...}, ...}
_CACHE_MTIME = 0             # mtime of the file we last parsed
_LAST_HASH = None            # MD5 hash of source file content
_LAST_COPY_ATTEMPT = 0       # timestamp of last failed copy attempt
_COPY_COOLDOWN = 3.0         # seconds between copy attempts when locked


# ── Shadow Copy Management ───────────────────────────────────────────────────

def _hash_file(filepath, blocksize=65536):
    """MD5 hash of file content. Returns None if file unreadable."""
    hasher = hashlib.md5()
    try:
        with open(filepath, 'rb') as f:
            while chunk := f.read(blocksize):
                hasher.update(chunk)
        return hasher.hexdigest()
    except (OSError, PermissionError):
        return None


def _try_copy_shadow():
    """
    Copy source to shadow if content actually changed (hash-based).
    Handles Excel-locked files gracefully.
    
    Returns (success: bool, used_shadow: bool, message: str)
        success     — operation completed (shadow is usable)
        used_shadow — True if we served from existing shadow (no copy made)
        message     — human-readable status for logging
    """
    global _LAST_HASH, _LAST_COPY_ATTEMPT
    
    source = get_source_path()
    shadow = get_shadow_path()
    
    # Source missing entirely
    if not os.path.exists(source):
        if os.path.exists(shadow):
            return True, True, "Source missing, using shadow"
        return False, False, "Source and shadow both missing"
    
    # Get current content hash
    current_hash = _hash_file(source)
    if current_hash is None:
        # Can't read source (locked) — use shadow if available
        if os.path.exists(shadow):
            return True, True, "Source locked, using shadow"
        return False, False, "Source locked, no shadow available"
    
    # Hash unchanged — no need to copy regardless of mtime
    if current_hash == _LAST_HASH and os.path.exists(shadow):
        return True, True, "Hash unchanged, using shadow"
    
    # Hash changed — need to update shadow
    now = time.time()
    _LAST_COPY_ATTEMPT = now
    
    try:
        shutil.copy2(source, shadow)
        _LAST_HASH = current_hash
        print(f"[ExcelStylesLoader] Shadow updated (hash: {current_hash[:8]}...)")
        return True, False, "Shadow updated from source"
        
    except (PermissionError, OSError) as e:
        # Locked — use existing shadow if we have one
        if os.path.exists(shadow):
            # Don't update _LAST_HASH — we'll try again next time
            return True, True, f"Source locked, using stale shadow ({e})"
        return False, False, f"Source locked, no shadow ({e})"


# ── Excel Parsing ────────────────────────────────────────────────────────────

def load_excel_data(force=False):
    """
    Parse Styles.xlsx via shadow copy.
    Call with force=True to bypass cache (user-triggered refresh).
    """
    global _EXCEL_CACHE, _CACHE_MTIME
    
    # Try to refresh shadow first
    success, used_shadow, msg = _try_copy_shadow()
    if not success:
        _EXCEL_CACHE = {"Error: " + msg: {"Check file": ["", ""]}}
        return _EXCEL_CACHE
    
    # Determine which file to read
    shadow = get_shadow_path()
    filepath = shadow if os.path.exists(shadow) else get_source_path()
    
    if not os.path.exists(filepath):
        _EXCEL_CACHE = {"Error: No Excel file available": {"": ["", ""]}}
        return _EXCEL_CACHE
    
    # Check if we need to reparse
    current_mtime = os.path.getmtime(filepath)
    if not force and _EXCEL_CACHE is not None and current_mtime == _CACHE_MTIME:
        return _EXCEL_CACHE
    
    data = {}
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            data[sheet_name] = {}
            
            rows = list(sheet.iter_rows(values_only=True))
            if not rows or len(rows) <= 1:
                continue
            
            # Robust header detection
            raw_headers = [str(c).strip() if c is not None else "" for c in rows[0]]
            headers = [h.lower() for h in raw_headers]
            
            def find_col(*candidates, default=0):
                for cand in candidates:
                    cand = cand.lower()
                    if cand in headers:
                        return headers.index(cand)
                    for i, h in enumerate(headers):
                        if cand.replace("_", " ") == h.replace("_", " "):
                            return i
                return default
            
            name_idx = find_col("name", "style", "title", default=0)
            pos_idx  = find_col("prompt", "positive", "positive prompt", default=1)
            neg_idx  = find_col("negative_prompt", "negative prompt", "neg", default=2)
            
            for row in rows[1:]:
                if not row or row[name_idx] is None:
                    continue
                
                style_name = str(row[name_idx]).strip()
                if not style_name:
                    continue
                
                pos_val = str(row[pos_idx]).strip() if len(row) > pos_idx and row[pos_idx] is not None else ""
                neg_val = str(row[neg_idx]).strip() if len(row) > neg_idx and row[neg_idx] is not None else ""
                
                data[sheet_name][style_name] = [pos_val, neg_val]
        
        _EXCEL_CACHE = data
        _CACHE_MTIME = current_mtime
        return data
        
    except Exception as e:
        print(f"[ExcelStylesLoader] Error reading workbook: {e}")
        _EXCEL_CACHE = {f"Error: {str(e)}": {"Check file formatting": ["", ""]}}
        return _EXCEL_CACHE


# ── HTTP Endpoints ───────────────────────────────────────────────────────────

@PromptServer.instance.routes.get("/custom_nodes/excel_styles_loader/get_tabs")
async def get_tabs_endpoint(request):
    """Return all sheet names (tabs). ?force=1 to bypass cache."""
    force = request.rel_url.query.get("force", "0") == "1"
    data = load_excel_data(force=force)
    tabs = [t for t in data.keys() if not t.startswith("Error:")]
    return web.json_response({
        "tabs": tabs,
        "error": next((k for k in data.keys() if k.startswith("Error:")), None)
    })


@PromptServer.instance.routes.get("/custom_nodes/excel_styles_loader/get_styles")
async def get_styles_endpoint(request):
    """Return style names for a given tab. ?force=1 to bypass cache."""
    tab_name = request.rel_url.query.get("tab_name", "")
    force = request.rel_url.query.get("force", "0") == "1"
    data = load_excel_data(force=force)
    
    styles = list(data.get(tab_name, {}).keys()) if tab_name in data else []
    return web.json_response({
        "styles": styles,
        "tab_exists": tab_name in data
    })


# ── ComfyUI Node ─────────────────────────────────────────────────────────────

class ExcelStylesLoader:
    @classmethod
    def INPUT_TYPES(cls):
        # Prime cache so we have real tab names on first node creation
        data = load_excel_data()
        tab_choices = [t for t in data.keys() if not t.startswith("Error:")]
        
        if not tab_choices:
            tab_choices = ["No Tabs Found"]
        
        return {
            "required": {
                "tab_name": (tab_choices,),
                "style_name": (["Select a Tab"],),
            }
        }

    RETURN_TYPES = ("STRING", "STRING")
    RETURN_NAMES = ("positive prompt", "negative prompt")
    FUNCTION = "execute"
    CATEGORY = "loaders"

    def execute(self, tab_name, style_name):
        data = load_excel_data()
        
        # Exact match
        if tab_name in data and style_name in data[tab_name]:
            return tuple(data[tab_name][style_name])
        
        # Tab exists but style missing — fallback to first style in tab
        if tab_name in data and data[tab_name]:
            first_key = next(iter(data[tab_name]))
            return tuple(data[tab_name][first_key])
        
        # Complete fallback
        return ("", "")

    @classmethod
    def VALIDATE_INPUTS(cls, tab_name, style_name, **kwargs):
        """
        Allow any values — the frontend sends dynamically populated strings
        that may not match the original INPUT_TYPES list. Actual validation
        happens in execute() with graceful fallbacks.
        """
        return True


NODE_CLASS_MAPPINGS = {"ExcelStylesLoader": ExcelStylesLoader}
NODE_DISPLAY_NAME_MAPPINGS = {"ExcelStylesLoader": "Load Styles from Excel (.xlsx)"}